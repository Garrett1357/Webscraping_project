import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()


ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman',size=24,italic=False,bold=True)

headerfont = Font(name='Times New Roman',size=24,italic=False,bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

#unmerge
#ws.unmerge_cells('A1:B1')



ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=SUM(B2:B7)'




#change the column width
ws.column_dimensions['A'].width = 25



# Display the grand total and average of 'amt sold' and 'total'
# at the bottom of the list along with appropriate details

#NEW
write_ws = wb['Second Sheet']
read_wb = xl.load_workbook("ProduceReport.xlsx")
read_ws = read_wb['ProduceReport']

for row in read_ws:
    ls = [i.value for i in row]
    write_ws.append(ls)

max_row = write_ws.max_row

write_ws.cell(max_row+2,1).value = "Totals:"
write_ws.cell(max_row+2,3).value = f"=SUM(C1:C{max_row})"
write_ws.cell(max_row+2,4).value = f"=SUM(D1:D{max_row})"


write_ws.cell(max_row+3,1).value = "Averages:"
write_ws.cell(max_row+3,3).value = f"=AVERAGE(C1:C{max_row})"
write_ws.cell(max_row+3,4).value = f"=AVERAGE(D1:D{max_row})"

for cell in write_ws["C:C"]:
    cell.number_format = '#,##0'

for cell in write_ws["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

#NEW
wb.save('pythontoexcel.xlsx')



write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook("ProduceReport.xlsx")
read_ws = read_wb['ProduceReport']

